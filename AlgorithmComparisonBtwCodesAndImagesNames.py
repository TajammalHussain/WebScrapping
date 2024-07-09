import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define file paths
excel_path = r'C:\Users\Tajammal\Desktop\MRF Brands upload\Wavin Full Codes List\Wavin Full Codes List.xlsx'
images_dir = r'C:\Users\Tajammal\Desktop\MRF Brands upload\Wavin images\DownloadedImages'
output_excel_path = r'C:\Users\Tajammal\Desktop\MRF Brands upload\Wavin Full Codes List\Wavin Full Codes List Output2.xlsx'

# Read the Excel file
df = pd.read_excel(excel_path)

# Extract product codes
product_codes = df['Product Code'].astype(str).tolist()

# List all image files in the directory
image_files = [os.path.splitext(f)[0] for f in os.listdir(images_dir) if f.endswith('.png')]

# Compare and find missing product codes
missing_codes = [code for code in product_codes if code not in image_files]

# Add a new column to the DataFrame indicating if the image is missing
df['Image Missing'] = df['Product Code'].astype(str).apply(lambda x: 'Yes' if x in missing_codes else 'No')

# Save the DataFrame to a new Excel file
df.to_excel(output_excel_path, index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(output_excel_path)
ws = wb.active

# Define a fill for highlighting
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Highlight the rows with missing images
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    if row[-1].value == 'Yes':  # Assuming 'Image Missing' is the last column
        for cell in row:
            cell.fill = highlight_fill

# Save the workbook with highlights
wb.save(output_excel_path)

print(f"Output file saved to {output_excel_path}")
